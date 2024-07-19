import openpyxl
from collections import defaultdict
import os
import pandas as pd
from app.models import Task
from app.database import db
from sqlalchemy.exc import SQLAlchemyError

def load_worker_data(file_path='data/worker_data.xlsx'):
    full_path = os.path.join(os.path.dirname(__file__), file_path)
    wb = openpyxl.load_workbook(full_path)
    sheet = wb.active
    
    supervisors = defaultdict(list)
    workers = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        supervisor, worker_name, worker_number, specialty, gender = row
        
        supervisors[supervisor].append(worker_name)
        workers[worker_name] = {
            'name': worker_name,
            'number': worker_number,
            'specialty': specialty.upper() if specialty else 'Not specified',
            'supervisor': supervisor,
            'gender': gender
        }
    
    # print("Loaded worker data:")
    # print("Supervisors:", dict(supervisors))
    # print("Workers:", workers)
    return dict(supervisors), workers

def load_project_data(file_path='data/project_data.xlsx'):
    full_path = os.path.join(os.path.dirname(__file__), file_path)
    wb = openpyxl.load_workbook(full_path)
    sheet = wb.active
    
    projects = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        project_name, total_houses, num_modulos = row
        projects[project_name] = {
            'total_houses': total_houses,
            'num_modulos': num_modulos
        }
    
    # print("Loaded project data:", projects)
    return projects

def load_activity_data(file_path='data/activity_data.xlsx'):
    full_path = os.path.join(os.path.dirname(__file__), file_path)
    wb = openpyxl.load_workbook(full_path)
    sheet = wb.active
    
    activities = defaultdict(list)
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        specialty, activity = row
        activities[specialty.upper()].append(activity)
    
    # print("Loaded activity data:", dict(activities))
    return dict(activities)

def get_active_tasks():
    try:
        return Task.query.filter_by(status='en proceso').all()
    except SQLAlchemyError as e:
        print(f"Error retrieving active tasks: {str(e)}")
        return []

def get_paused_tasks():
    try:
        return Task.query.filter_by(status='Paused').all()
    except SQLAlchemyError as e:
        print(f"Error retrieving paused tasks: {str(e)}")
        return []

def get_finished_tasks(start_date, end_date):
    try:
        return Task.query.filter(
            Task.status == 'Finished',
            Task.end_time >= start_date,
            Task.end_time <= end_date
        ).all()
    except SQLAlchemyError as e:
        print(f"Error retrieving finished tasks: {str(e)}")
        return []

def sync_task_data_to_excel(filename):
    try:
        tasks = Task.query.all()
        df = pd.DataFrame([
            {
                'worker_number': task.worker_number,
                'worker_name': task.worker_name,
                'project': task.project,
                'house': task.house,
                'module': task.module,
                'activity': task.activity,
                'start_time': task.start_time,
                'end_time': task.end_time,
                'status': task.status,
                'comment': task.comment
            } for task in tasks
        ])
        df.to_excel(filename, index=False)
        print(f"Task data synced to {filename}")
    except Exception as e:
        print(f"Error syncing task data to Excel: {str(e)}")

def get_task_statistics():
    try:
        total_tasks = Task.query.count()
        active_tasks = Task.query.filter_by(status='en proceso').count()
        finished_tasks = Task.query.filter_by(status='Finished').count()
        return {
            'total': total_tasks,
            'active': active_tasks,
            'finished': finished_tasks
        }
    except SQLAlchemyError as e:
        print(f"Error getting task statistics: {str(e)}")
        return {'total': 0, 'active': 0, 'finished': 0}

def safe_db_operation(operation):
    try:
        result = operation()
        db.session.commit()
        return result
    except SQLAlchemyError as e:
        db.session.rollback()
        print(f"Database error: {str(e)}")
        return None

def add_new_task(worker_number, worker_name, project, house, module, activity):
    return safe_db_operation(lambda: Task.add_task(worker_number, worker_name, project, house, module, activity))

# Test the functions
if __name__ == "__main__":
    supervisors, workers = load_worker_data()
    projects = load_project_data()
    activities = load_activity_data()
    print("Active tasks:", get_active_tasks())
    print("Task statistics:", get_task_statistics())
