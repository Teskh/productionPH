from app.utils import format_timestamp
from filelock import FileLock
import openpyxl
from datetime import datetime
import os

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'production_data.xlsx')
EXCEL_LOCK_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data', 'production_data.xlsx.lock')

class Task:
    @staticmethod
    def get_active_tasks(worker_number):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            active_tasks = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 10 and str(row[2]) == str(worker_number) and row[10] in ['en proceso', 'Paused']:
                    task = {
                        'task_id': row[0],
                        'start_time': row[1],
                        'worker_number': row[2],
                        'user': row[4],
                        'project': row[6],
                        'house_number': row[7],
                        'n_modulo': row[8],
                        'activity': row[9],
                        'status': row[10],
                        'station_i': row[11] if len(row) > 11 else '',
                        'station_f': row[12] if len(row) > 12 else '',
                        'line': row[13] if len(row) > 13 else '',
                    }
                    if len(row) > 13:
                        task['pause_1_time'] = row[13]
                        task['pause_1_reason'] = row[14] if len(row) > 14 else ''
                    if len(row) > 15:
                        task['resume_1_time'] = row[15]
                    if len(row) > 16:
                        task['pause_2_time'] = row[16]
                        task['pause_2_reason'] = row[17] if len(row) > 17 else ''
                    if len(row) > 18:
                        task['resume_2_time'] = row[18]
                    active_tasks.append(task)
            return active_tasks

    @staticmethod
    def get_all_active_tasks():
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            active_tasks = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) > 10 and row[10] == 'en proceso':
                    task = {
                        'task_id': row[0],
                        'worker_number': row[2],
                    }
                    active_tasks.append(task)
            return active_tasks

    @staticmethod
    def get_active_task(worker_number):
        tasks = Task.get_active_tasks(worker_number)
        return next((task for task in tasks if task['status'] == 'en proceso'), None)

    @staticmethod
    def add_task(task_data):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            new_row = [
                task_data.get('task_id', ''),
                format_timestamp(),
                task_data.get('worker_number', ''),
                task_data.get('supervisor', ''),
                task_data.get('user', ''),
                task_data.get('specialty', ''),
                task_data.get('project', ''),
                task_data.get('house_number', ''),
                task_data.get('n_modulo', ''),
                task_data.get('activity', ''),
                'en proceso',  # Default status when starting a new task
                task_data.get('station_i', ''),
                '',  # station_f (empty initially)
                task_data.get('line', ''),
                '', '', '', '', '', '',  # Pause and Resume columns
                ''  # End Timestamp
            ]
            ws.append(new_row)
            wb.save(EXCEL_FILE)
        return True  # Indicate successful task addition

    @staticmethod
    def update_task(task_id, new_status, timestamp=None, pause_reason=None, station=None):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=False):
                if row[0].value == task_id:
                    # Ensure the row has enough cells
                    while len(row) < 20:
                        ws.cell(row=row[0].row, column=len(row)+1, value='')
                    
                    row[10].value = new_status
                    if new_status == 'Paused':
                        if not row[14].value:  # First pause
                            row[14].value = timestamp
                            row[15].value = pause_reason
                        elif not row[17].value:  # Second pause
                            row[17].value = timestamp
                            row[18].value = pause_reason
                    elif new_status == 'en proceso':
                        if row[14].value and not row[16].value:  # First resume
                            row[16].value = timestamp
                        elif row[17].value and not row[19].value:  # Second resume
                            row[19].value = timestamp
                    elif new_status == 'Finished':
                        row[20].value = timestamp
                        row[12].value = station  # Update station_f when task is finished
                    break
            wb.save(EXCEL_FILE)

    @staticmethod
    def get_user_tasks(worker_number):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            user_tasks = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if str(row[2]) == str(worker_number):
                    task = {
                        'task_id': row[0],
                        'start_time': row[1],
                        'worker_number': row[2],
                        'user': row[4],
                        'project': row[6],
                        'house_number': row[7],
                        'n_modulo': row[8],
                        'activity': row[9],
                        'status': row[10],
                    }
                    user_tasks.append(task)
            return user_tasks

    @staticmethod
    def get_related_active_tasks(project, house_number, n_modulo, activity):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            related_tasks = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if (row[6] == project and 
                    str(row[7]) == str(house_number) and 
                    str(row[8]) == str(n_modulo) and 
                    row[9] == activity and 
                    row[10] == 'en proceso'):
                    related_tasks.append({
                        'task_id': row[0],
                        'worker_number': row[2],
                        'status': row[10],
                    })
            return related_tasks

    @staticmethod
    def finish_related_tasks(project, house_number, n_modulo, activity, timestamp, station):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=False):
                if (row[6].value == project and 
                    str(row[7].value) == str(house_number) and 
                    str(row[8].value) == str(n_modulo) and 
                    row[9].value == activity and 
                    (row[10].value == 'en proceso' or row[10].value == 'Paused')):
                    row[10].value = 'Finished'
                    row[20].value = timestamp
                    row[12].value = station
            wb.save(EXCEL_FILE)

    @staticmethod
    def get_task_by_id(task_id):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == task_id:
                    return {
                        'task_id': row[0],
                        'start_time': row[1],
                        'worker_number': row[2],
                        'user': row[4],
                        'project': row[6],
                        'house_number': row[7],
                        'n_modulo': row[8],
                        'activity': row[9],
                        'status': row[10],
                        'station_i': row[11],
                        'station_f': row[12],
                    }
            return None

    @staticmethod
    def get_finished_task(project, house_number, n_modulo, activity):
        lock = FileLock(EXCEL_LOCK_FILE)
        with lock:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if (row[6] == project and 
                    str(row[7]) == str(house_number) and 
                    str(row[8]) == str(n_modulo) and 
                    row[9] == activity and 
                    row[10] == 'Finished'):
                    return {
                        'task_id': row[0],
                        'start_time': row[1],
                        'worker_number': row[2],
                        'user': row[4],
                        'project': row[6],
                        'house_number': row[7],
                        'n_modulo': row[8],
                        'activity': row[9],
                        'status': row[10],
                        'station_i': row[11],
                        'station_f': row[12],
                    }
            return None
