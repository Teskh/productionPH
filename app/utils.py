from datetime import datetime
import openpyxl
import os

EXCEL_FILE = 'data/production_data.xlsx'
EXPECTED_COLUMNS = [
    'Task ID', 'Start Timestamp', 'Worker Number', 'Supervisor', 'User', 'Specialty', 'Project', 
    'House Number', 'Nº modulo', 'Activity', 'Status', 'Station', 'Line',
    'Pause 1 Timestamp', 'Pause 1 Reason', 'Resume 1 Timestamp',
    'Pause 2 Timestamp', 'Pause 2 Reason', 'Resume 2 Timestamp',
    'End Timestamp'
]

def format_timestamp():
    return datetime.now().strftime('%Y-%m-%d %H:%M')

def init_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(EXPECTED_COLUMNS)
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws[1][0].value != EXPECTED_COLUMNS[0] or len(ws[1]) < len(EXPECTED_COLUMNS):
            for col, header in enumerate(EXPECTED_COLUMNS, start=1):
                if col > len(ws[1]) or ws.cell(row=1, column=col).value != header:
                    ws.cell(row=1, column=col, value=header)
    wb.save(EXCEL_FILE)
